from django.db import models
from django.http import HttpResponse

from rest_framework import viewsets, filters, status as drf_status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, JSONParser
from rest_framework.response import Response
from rest_framework.views import APIView

from django_filters.rest_framework import DjangoFilterBackend, FilterSet, DateFilter

import openpyxl
from openpyxl import Workbook

from .models import Product, Supply, Shrinkage, ReturnSupplier
from suppliers.models import Supplier
from .serializers import (
    ProductSerializer,
    SupplySerializer,
    ShrinkageSerializer,
    ReturnSupplierSerializer
)
from users.pagination import CustomPagination
from .filters import ProductFilter


# -------------------------------
# PRODUCTOS
# -------------------------------
class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ProductFilter
    search_fields = ['name', 'category']
    ordering_fields = ['name', 'price_clp', 'stock']
    ordering = ['name']
    parser_classes = [MultiPartParser, JSONParser]

    @action(detail=False, methods=['get'], url_path='low-stock')
    def low_stock_products(self, request):
        low_stock = Product.objects.filter(stock__lt=models.F('min_stock'))
        page = self.paginate_queryset(low_stock)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(low_stock, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='bulk-upload-excel')
    def bulk_upload_excel(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "Archivo no proporcionado."}, status=drf_status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception:
            return Response({"error": "Archivo inválido o corrupto."}, status=drf_status.HTTP_400_BAD_REQUEST)

        products_data = []
        errores = []

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Verifica si la fila está completamente vacía
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue  # Ignora filas completamente vacías

            if not row or len(row) < 7:
                errores.append(f"Fila {i}: Datos incompletos.")
                continue

            name, price_clp, iva, stock, min_stock, category, supplier_name = row

            if not all([name, price_clp, iva, stock, min_stock, category]):
                errores.append(f"Fila {i}: Uno o más campos requeridos están vacíos.")
                continue

            # Convertir IVA
            iva = True if str(iva).strip().upper() in ["VERDADERO", "TRUE", "1"] else False

            # Usar "Sin proveedor" si está vacío
            if not supplier_name:
                supplier_name = "Sin proveedor"

            supplier_obj = Supplier.objects.filter(name__iexact=supplier_name.strip()).first()

            if not supplier_obj:
                supplier_obj = Supplier.objects.create(name=supplier_name.strip())

            products_data.append({
                "name": name,
                "price_clp": price_clp,
                "iva": iva,
                "stock": stock,
                "min_stock": min_stock,
                "category": category,
                "supplier": str(supplier_obj.id)
            })

        if errores:
            return Response({"errores": errores}, status=drf_status.HTTP_400_BAD_REQUEST)

        serializer = ProductSerializer(data=products_data, many=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Productos cargados exitosamente."}, status=drf_status.HTTP_201_CREATED)

        return Response(serializer.errors, status=drf_status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='excel-template')
    def download_template(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Productos"
        ws.append(["name", "price_clp", "iva", "stock", "min_stock", "category", "supplier"])
        ws.append(["Polera básica", 7990, True, 50, 10, "Ropa", "Proveedor A"])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=plantilla_productos.xlsx'
        wb.save(response)
        return response


# -------------------------------
# INSUMOS
# -------------------------------
class SupplyViewSet(viewsets.ModelViewSet):
    queryset = Supply.objects.all()
    serializer_class = SupplySerializer
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category']
    search_fields = ['name', 'category']
    ordering_fields = ['name', 'category', 'stock']
    ordering = ['name']
    parser_classes = [MultiPartParser, JSONParser]

    @action(detail=False, methods=['get'], url_path='low-stock')
    def low_stock_supplies(self, request):
        low_stock = Supply.objects.filter(stock__lt=models.F('min_stock'))
        page = self.paginate_queryset(low_stock)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(low_stock, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='bulk-upload-excel')
    def bulk_upload_excel(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "Archivo no proporcionado."}, status=drf_status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception:
            return Response({"error": "Archivo inválido o corrupto."}, status=drf_status.HTTP_400_BAD_REQUEST)

        supplies_data = []
        errores = []

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            if len(row) < 4:
                errores.append(f"Fila {i}: Datos incompletos.")
                continue

            name, category, stock, min_stock = row

            if not all([name, category, stock, min_stock]):
                errores.append(f"Fila {i}: Uno o más campos requeridos están vacíos.")
                continue

            supplies_data.append({
                "name": name,
                "category": category,
                "stock": stock,
                "min_stock": min_stock,
            })

        if errores:
            return Response({"errores": errores}, status=drf_status.HTTP_400_BAD_REQUEST)

        serializer = SupplySerializer(data=supplies_data, many=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Insumos cargados exitosamente."}, status=drf_status.HTTP_201_CREATED)

        return Response(serializer.errors, status=drf_status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='excel-template')
    def download_template(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Insumos"
        ws.append(["name", "category", "stock", "min_stock"])
        ws.append(["Tela algodón", "Textil", 100, 20])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=plantilla_insumos.xlsx'
        wb.save(response)
        return response

# -------------------------------
# MERMAS
# -------------------------------
class ShrinkageViewSet(viewsets.ModelViewSet):
    queryset = Shrinkage.objects.all()
    serializer_class = ShrinkageSerializer
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category']
    search_fields = ['product', 'category', 'observation']
    ordering_fields = ['product', 'category', 'quantity', 'price']
    ordering = ['product']
    parser_classes = [JSONParser, MultiPartParser]

    @action(detail=False, methods=['post'], url_path='bulk-upload-excel')
    def bulk_upload_excel(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "Archivo no proporcionado."}, status=drf_status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception:
            return Response({"error": "Archivo inválido o corrupto."}, status=drf_status.HTTP_400_BAD_REQUEST)

        shrinkages_data = []
        errores = []

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            if len(row) < 4:
                errores.append(f"Fila {i}: Datos incompletos. Se requieren al menos 4 columnas.")
                continue

            product, price, quantity, category = row[:4]
            observation = row[4] if len(row) > 4 else ""

            if not all([product, price, quantity, category]):
                errores.append(f"Fila {i}: Uno o más campos requeridos están vacíos.")
                continue

            try:
                price = float(price)
                quantity = int(quantity)
            except ValueError:
                errores.append(f"Fila {i}: Precio o cantidad con formato incorrecto.")
                continue

            shrinkages_data.append({
                "product": str(product),
                "price": price,
                "quantity": quantity,
                "category": str(category),
                "observation": observation or "",
            })

        if errores:
            return Response({"errores": errores}, status=drf_status.HTTP_400_BAD_REQUEST)

        serializer = ShrinkageSerializer(data=shrinkages_data, many=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Mermas cargadas exitosamente."}, status=drf_status.HTTP_201_CREATED)

        return Response(serializer.errors, status=drf_status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='excel-template')
    def download_template(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Mermas"
        ws.append(["product", "price", "quantity", "category", "observation"])
        ws.append(["Polera básica", 3990, 2, "Ropa", "Rota en costura"])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=plantilla_mermas.xlsx'
        wb.save(response)
        return response

# -----------------------
# Devoluciones a proveedores
# -----------------------

class ReturnSupplierFilter(FilterSet):
    start_date = DateFilter(field_name="return_date", lookup_expr='gte')
    end_date = DateFilter(field_name="return_date", lookup_expr='lte')

    class Meta:
        model = ReturnSupplier
        fields = ['supplier', 'product', 'status', 'start_date', 'end_date']

class ReturnSupplierViewSet(viewsets.ModelViewSet):
    queryset = ReturnSupplier.objects.all()
    serializer_class = ReturnSupplierSerializer
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ReturnSupplierFilter
    search_fields = ['reason', 'purchase_number']
    ordering_fields = ['return_date', 'purchase_date', 'quantity']
    ordering = ['-return_date']

    @action(detail=True, methods=["patch"], url_path="resolve")
    def mark_as_resolved(self, request, pk=None):
        instance = self.get_object()
        if instance.status == "Resuelto":
            return Response(
                {"message": "Esta devolución ya está marcada como Resuelto."},
                status=drf_status.HTTP_400_BAD_REQUEST
            )
        instance.status = "Resuelto"
        instance.save()
        return Response(
            {"message": "Devolución marcada como Resuelto."},
            status=drf_status.HTTP_200_OK
        )

# -----------------------
# Metricas para Dashboard inventario
# -----------------------

class InventoryMetricsAPIView(APIView):

    def get(self, request):
        amount_products = Product.objects.count()
        amount_supplies = Supply.objects.count()
        amount_shrinkages = Shrinkage.objects.count()
        low_stock_products = Product.objects.filter(stock__lt=models.F('min_stock'))

        # Mermas recientes (últimos 5 registros)
        recent_shrinkages = Shrinkage.objects.order_by('-id')[:5]
        recent_shrinkages_serialized = [
            {
                "product_name": s.product,
                "quantity": s.quantity,
                "reason": s.observation,
                "created_at": s.created_at.isoformat() if s.created_at else None
            } for s in recent_shrinkages
        ]

        # Datos para gráfico
        low_stock_chart_data = [
            {
                "product": p.name,
                "stock": p.stock
            } for p in low_stock_products[:10]
        ]

        return Response({
            "amount_products": amount_products,
            "amount_supplies": amount_supplies,
            "amount_shrinkages": amount_shrinkages,
            "low_stock_products_count": low_stock_products.count(),
            "recent_shrinkages": recent_shrinkages_serialized,
            "low_stock_chart_data": low_stock_chart_data
        })
